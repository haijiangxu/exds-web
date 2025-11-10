from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
import pandas as pd
import io
from datetime import datetime
from urllib.parse import quote
from webapp.models.contract import Contract, ContractCreate, ContractListResponse, calculate_contract_status
from webapp.services.contract_service import ContractService
from webapp.tools.mongo import DATABASE
from webapp.tools.security import get_current_active_user, User
from webapp.utils.excel_handler import ExcelReader, DataValidator, ContractDataTransformer

router = APIRouter(prefix="/retail-contracts", tags=["Retail Contracts"])


@router.post("", response_model=dict, status_code=status.HTTP_201_CREATED)
async def create_contract(
    contract: ContractCreate,
    current_user: User = Depends(get_current_active_user)
):
    """创建新合同"""
    service = ContractService(DATABASE)
    try:
        result = service.create_contract(
            contract_data=contract.model_dump(exclude_unset=True),
            operator=current_user.username
        )
        return result
    except ValueError as e:
        error_msg = str(e)
        if "不存在" in error_msg or "无效" in error_msg or "状态不是" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )


@router.get("", response_model=ContractListResponse)
async def list_contracts(
    contract_name: Optional[str] = Query(None, description="合同名称（模糊搜索）"),
    package_name: Optional[str] = Query(None, description="套餐名称（模糊搜索）"),
    customer_name: Optional[str] = Query(None, description="客户名称（模糊搜索）"),
    status: Optional[str] = Query(None, description="合同状态（pending/active/expired）"),
    purchase_start_month: Optional[str] = Query(None, description="购电开始月份筛选（yyyy-MM）"),
    purchase_end_month: Optional[str] = Query(None, description="购电结束月份筛选（yyyy-MM）"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页大小"),
    current_user: User = Depends(get_current_active_user)
):
    """
    获取合同列表

    支持筛选：
    - contract_name: 合同名称（模糊搜索）
    - package_name: 套餐名称（模糊搜索）
    - customer_name: 客户名称（模糊搜索）
    - status: 合同状态（pending/active/expired）

    支持分页：
    - page: 页码（从1开始）
    - page_size: 每页数量
    """
    service = ContractService(DATABASE)
    result = service.list_contracts(
        filters={
            "contract_name": contract_name,
            "package_name": package_name,
            "customer_name": customer_name,
            "status": status,
            "purchase_start_month": purchase_start_month,
            "purchase_end_month": purchase_end_month
        },
        page=page,
        page_size=page_size
    )
    return result


@router.post("/import", summary="导入合同数据")
async def import_contracts(
    file: UploadFile = File(..., description="交易中心平台下载的Excel文件"),
    current_user: User = Depends(get_current_active_user)
):
    """
    批量导入合同数据（同步方案）

    Excel文件格式要求：
    - 必需列：套餐, 购买用户, 购买电量, 购买时间-开始, 购买时间-结束
    - 忽略列：序号, 代理销售费模型, 签章状态
    - 套餐名称必须存在于系统中且状态为已生效
    - 客户名称必须存在于系统中且状态为正常
    - 日期格式：YYYY-MM 或 YYYY-MM-DD
    - 购买电量必须大于0
    - 购买结束月份必须 >= 购买开始月份

    返回：
    - total: 总行数
    - success: 成功导入的数量
    - failed: 失败的数量
    - errors: 错误详情列表
    """
    try:
        # 1. 读取Excel文件
        contents = await file.read()

        # 2. 初始化处理组件
        excel_reader = ExcelReader()
        validator = DataValidator(DATABASE)
        transformer = ContractDataTransformer(DATABASE)

        # 3. 读取并验证文件结构
        df = excel_reader.read_excel_file(contents)
        excel_reader.validate_excel_structure(df)

        # 4. 逐行校验和导入
        success_count = 0
        failed_count = 0
        errors = []

        for index, row in df.iterrows():
            row_number = index + 2  # Excel从1开始，且有表头

            try:
                # 解析行数据
                row_data = excel_reader.parse_row_data(row, row_number)

                # 执行多层校验
                validation_errors = []
                validation_errors.extend(validator.validate_required_fields(row_data))
                validation_errors.extend(validator.validate_related_data(row_data))
                validation_errors.extend(validator.validate_business_rules(row_data))

                if validation_errors:
                    errors.extend(validation_errors)
                    failed_count += 1
                    continue

                # 转换数据格式
                contract_data = transformer.transform_row_to_contract(row_data, current_user.username)

                # 检查合同唯一性
                uniqueness_errors = validator.validate_contract_uniqueness(
                    contract_data['customer_id'],
                    contract_data['purchase_start_month'],
                    contract_data['purchase_end_month'],
                    row_number
                )

                if uniqueness_errors:
                    errors.extend(uniqueness_errors)
                    failed_count += 1
                    continue

                # 插入数据库
                DATABASE.retail_contracts.insert_one(contract_data)
                success_count += 1

            except Exception as e:
                errors.append({
                    'row': row_number,
                    'field': 'general',
                    'value': None,
                    'message': str(e),
                    'suggestion': '请检查该行数据的完整性和格式'
                })
                failed_count += 1

        # 5. 返回结果
        return {
            "total": len(df),
            "success": success_count,
            "failed": failed_count,
            "errors": errors
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"导入失败：{str(e)}")


@router.get("/export", summary="导出合同数据")
async def export_contracts(
    package_name: Optional[str] = Query(None, description="套餐名称筛选"),
    customer_name: Optional[str] = Query(None, description="客户名称筛选"),
    status: Optional[str] = Query(None, description="合同状态筛选"),
    start_month: Optional[str] = Query(None, description="购电开始月份筛选(YYYY-MM)"),
    end_month: Optional[str] = Query(None, description="购电结束月份筛选(YYYY-MM)"),
    current_user: User = Depends(get_current_active_user)
):
    """
    导出合同数据为Excel文件

    支持筛选：
    - package_name: 套餐名称（模糊搜索）
    - customer_name: 客户名称（模糊搜索）
    - status: 合同状态（pending/active/expired）
    - start_month: 购电开始月份筛选
    - end_month: 购电结束月份筛选
    """
    try:
        # 1. 构建查询条件
        query = {}
        if package_name:
            query["package_name"] = {"$regex": package_name, "$options": "i"}
        if customer_name:
            query["customer_name"] = {"$regex": customer_name, "$options": "i"}
        if start_month:
            query["purchase_start_month"] = {"$gte": start_month}
        if end_month:
            query["purchase_end_month"] = {"$lte": end_month}

        # 2. 查询数据
        contracts = list(DATABASE.retail_contracts.find(query).sort("created_at", -1))

        # 3. 计算虚拟状态并添加到数据中
        processed_contracts = []
        for contract in contracts:
            status_value = calculate_contract_status(
                contract.get("purchase_start_month"),
                contract.get("purchase_end_month")
            )

            # 应用状态筛选
            if status and status != "all":
                if status_value != status:
                    continue

            # 格式化数据
            formatted_contract = {
                '合同编号': str(contract.get('_id', '')),
                '合同名称': contract.get('contract_name', ''),
                '套餐名称': contract.get('package_name', ''),
                '购买用户': contract.get('customer_name', ''),
                '购买电量': contract.get('purchasing_electricity_quantity', 0),
                '购电开始月份': _format_date(contract.get('purchase_start_month'), 'date'),
                '购电结束月份': _format_date(contract.get('purchase_end_month'), 'date'),
                '合同状态': _format_status(status_value),
                '创建时间': _format_date(contract.get('created_at'), 'datetime'),
                '更新时间': _format_date(contract.get('updated_at'), 'datetime'),
            }

            processed_contracts.append(formatted_contract)

        # 4. 生成Excel文件
        excel_data = _generate_excel_file(processed_contracts, {
            'package_name': package_name,
            'customer_name': customer_name,
            'status': status,
            'start_month': start_month,
            'end_month': end_month
        })

        # 5. 返回文件流
        filename = f"零售合同数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # 对文件名进行 URL 编码以支持中文
        encoded_filename = quote(filename)

        return StreamingResponse(
            io.BytesIO(excel_data.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败：{str(e)}")


@router.get("/{contract_id}", response_model=dict)
async def get_contract(
    contract_id: str,
    current_user: User = Depends(get_current_active_user)
):
    """获取合同详情"""
    service = ContractService(DATABASE)
    try:
        result = service.get_contract_by_id(contract_id)
        return result
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e)
        )


@router.put("/{contract_id}", response_model=dict)
async def update_contract(
    contract_id: str,
    contract: ContractCreate,
    current_user: User = Depends(get_current_active_user)
):
    """更新合同（仅待生效状态）"""
    service = ContractService(DATABASE)
    try:
        result = service.update_contract(
            contract_id=contract_id,
            contract_data=contract.model_dump(exclude_unset=True),
            operator=current_user.username
        )
        return result
    except ValueError as e:
        error_msg = str(e)
        if "不存在" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=error_msg
            )
        elif "状态" in error_msg or "无效" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )


@router.delete("/{contract_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_contract(
    contract_id: str,
    current_user: User = Depends(get_current_active_user)
):
    """删除合同（仅待生效状态）"""
    service = ContractService(DATABASE)
    try:
        service.delete_contract(contract_id)
        return None  # 204 No Content
    except ValueError as e:
        error_msg = str(e)
        if "状态" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=error_msg
            )


def _format_date(date_value, format_type='date'):
    """格式化日期"""
    if date_value is None:
        return ""

    if isinstance(date_value, str):
        date_value = datetime.fromisoformat(date_value.replace('Z', '+00:00'))

    if format_type == 'date':
        return date_value.strftime("%Y-%m")
    elif format_type == 'datetime':
        return date_value.strftime("%Y-%m-%d %H:%M:%S")
    else:
        return str(date_value)


def _format_status(status):
    """格式化合同状态为中文"""
    status_map = {
        'pending': '待生效',
        'active': '生效',
        'expired': '已过期'
    }
    return status_map.get(status, status)


def _generate_excel_file(data, params):
    """生成Excel文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # 创建DataFrame
    if not data:
        # 创建空表格带列名
        columns = [
            '合同编号', '合同名称', '套餐名称', '购买用户', '购买电量',
            '购电开始月份', '购电结束月份', '合同状态', '创建时间', '更新时间'
        ]
        df = pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(data)

    # 创建Excel工作簿
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='合同数据')

        # 获取工作表
        worksheet = writer.sheets['合同数据']

        # 设置样式
        _format_excel_worksheet(worksheet, df, params)

    output.seek(0)
    return output


def _format_excel_worksheet(worksheet, df, params):
    """格式化Excel工作表"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 设置标题行样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 设置数据行样式
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')

    # 设置边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用标题行样式
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 应用数据行样式
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # 状态列颜色标记
            if cell.column_letter == 'H':  # 状态列
                if cell.value == '生效':
                    cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                elif cell.value == '已过期':
                    cell.fill = PatternFill(start_color='FFF2E8', end_color='FFF2E8', fill_type='solid')
                elif cell.value == '待生效':
                    cell.fill = PatternFill(start_color='E8F4FF', end_color='E8F4FF', fill_type='solid')

    # 自动调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 冻结首行
    worksheet.freeze_panes = 'A2'

    # 添加筛选器
    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

    # 添加页眉信息
    if not df.empty:
        # 在顶部添加筛选条件说明
        worksheet.insert_rows(1)
        worksheet.cell(row=1, column=1, value='导出条件：')

        filter_desc = _build_filter_description(params)
        worksheet.cell(row=1, column=2, value=filter_desc)

        # 合并单元格
        worksheet.merge_cells(f'B1:{get_column_letter(worksheet.max_column)}1')

        # 设置页眉样式
        header_cell = worksheet.cell(row=1, column=1)
        header_cell.font = Font(name='微软雅黑', size=10, bold=True)
        header_cell.alignment = Alignment(horizontal='left', vertical='center')


def _build_filter_description(params):
    """构建筛选条件描述"""
    descriptions = []

    if params.get('package_name'):
        descriptions.append(f"套餐名称：{params['package_name']}")

    if params.get('customer_name'):
        descriptions.append(f"客户名称：{params['customer_name']}")

    if params.get('status') and params['status'] != 'all':
        status_map = {'pending': '待生效', 'active': '生效', 'expired': '已过期'}
        descriptions.append(f"合同状态：{status_map.get(params['status'], params['status'])}")

    if params.get('start_month') or params.get('end_month'):
        start = params.get('start_month', '开始')
        end = params.get('end_month', '结束')
        descriptions.append(f"购电时间：{start} ~ {end}")

    return '；'.join(descriptions) if descriptions else '无筛选条件'
